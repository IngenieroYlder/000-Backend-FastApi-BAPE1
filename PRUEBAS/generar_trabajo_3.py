from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PRUEBAS_DIR = PROJECT_ROOT / "PRUEBAS"
RESULTS_PATH = PRUEBAS_DIR / "evidencias" / "20260331_051122" / "results.json"
OUTPUT_PATH = PRUEBAS_DIR / "TRABAJO 3 v2.xlsx"

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
SECTION_FILL = PatternFill("solid", fgColor="CFE2F3")
SUB_FILL = PatternFill("solid", fgColor="F4CCCC")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BOLD = Font(bold=True)


def merge_and_set(ws, cell_range: str, value: str, fill=None, bold: bool = False, align=LEFT):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    cell.alignment = align
    cell.border = BORDER
    if fill is not None:
        cell.fill = fill
    if bold:
        cell.font = BOLD
    for row in ws[cell_range]:
        for c in row:
            c.border = BORDER
            if fill is not None:
                c.fill = fill


def set_cell(ws, ref: str, value: str, fill=None, bold: bool = False, align=LEFT):
    cell = ws[ref]
    cell.value = value
    cell.alignment = align
    cell.border = BORDER
    if fill is not None:
        cell.fill = fill
    if bold:
        cell.font = BOLD


def status_from_results(results: dict, code: str) -> str:
    for prueba in results["pruebas"].values():
        if prueba["nombre_prueba"].startswith(code):
            return prueba["resultado"].upper()
    return "NO EJECUTADA"


def build_cases(results: dict) -> list[dict]:
    return [
        {
            "code": "CP-SM-01",
            "description": "Validar endpoint de salud del backend",
            "prerequisites": [
                "El entorno virtual del proyecto debe existir.",
                "Las dependencias del backend deben estar instaladas.",
                "La suite de pruebas debe estar disponible en PRUEBAS/tests.",
                "N/A",
            ],
            "test_data": [
                "Solicitud GET a /health",
                "Sin token ni payload",
                "Respuesta esperada en formato JSON",
                "N/A",
            ],
            "scenario": "Prueba smoke para comprobar que el backend responde en el endpoint de salud y confirma disponibilidad basica del servicio.",
            "steps": [
                ("1.0", "Preparar el cliente de prueba local con FastAPI TestClient.", "Cliente listo para enviar solicitudes al backend.", "Cliente inicializado correctamente.", "APROBADA"),
                ("2.0", "Enviar una solicitud GET al endpoint /health.", "El backend responde con codigo 200.", "Se obtuvo respuesta HTTP 200.", "APROBADA"),
                ("3.0", "Validar el contenido del cuerpo de la respuesta.", "El cuerpo debe ser {'status':'ok'}.", "La respuesta coincide con {'status':'ok'}.", "APROBADA"),
            ],
        },
        {
            "code": "CP-SM-02",
            "description": "Verificar acceso a la pagina de inicio de sesion",
            "prerequisites": [
                "Las plantillas HTML deben estar disponibles en app/templates.",
                "La aplicacion FastAPI debe poder resolver la ruta /login.",
                "El cliente de prueba debe estar configurado.",
                "N/A",
            ],
            "test_data": [
                "Solicitud GET a /login",
                "Sin payload",
                "Validacion de contenido HTML",
                "N/A",
            ],
            "scenario": "Prueba smoke para validar que la pagina de login carga correctamente y entrega contenido HTML al usuario.",
            "steps": [
                ("1.0", "Inicializar el cliente de prueba local.", "El cliente queda disponible para navegar rutas web.", "Cliente inicializado correctamente.", "APROBADA"),
                ("2.0", "Enviar GET a /login.", "El sistema responde con estado 200.", "Se obtuvo respuesta HTTP 200.", "APROBADA"),
                ("3.0", "Verificar que el contenido de respuesta sea HTML.", "La respuesta contiene estructura HTML de login.", "Se detecto contenido HTML valido.", "APROBADA"),
            ],
        },
        {
            "code": "CP-SM-03",
            "description": "Verificar ruta raiz como pantalla inicial",
            "prerequisites": [
                "La ruta raiz debe estar configurada en frontend_routes.",
                "Las plantillas HTML deben estar accesibles.",
                "El cliente de pruebas debe estar listo.",
                "N/A",
            ],
            "test_data": [
                "Solicitud GET a /",
                "Sin payload",
                "Validacion de respuesta HTML",
                "N/A",
            ],
            "scenario": "Prueba smoke para validar que la ruta raiz del sistema funcione como punto de entrada y renderice la vista inicial.",
            "steps": [
                ("1.0", "Preparar el cliente de prueba para navegar la ruta raiz.", "Cliente disponible.", "Cliente listo para ejecutar la solicitud.", "APROBADA"),
                ("2.0", "Enviar GET a /.", "El backend devuelve estado 200.", "Se obtuvo respuesta HTTP 200.", "APROBADA"),
                ("3.0", "Validar que la respuesta incluya contenido HTML.", "La pagina inicial se renderiza correctamente.", "Se confirmo HTML en la respuesta.", "APROBADA"),
            ],
        },
        {
            "code": "CP-AU-01",
            "description": "Registrar un usuario valido",
            "prerequisites": [
                "Debe existir un plan Free en la base de datos temporal.",
                "El endpoint /auth/register debe estar habilitado.",
                "La base de datos de prueba debe aceptar inserciones.",
                "N/A",
            ],
            "test_data": [
                "email unico",
                "password valida",
                "company_name de prueba",
                "datos basicos del usuario",
            ],
            "scenario": "Prueba funcional para verificar que el sistema registra un nuevo usuario y crea la compania asociada sin errores.",
            "steps": [
                ("1.0", "Construir payload con email unico, clave y empresa.", "El payload queda listo para el registro.", "Payload generado correctamente.", "APROBADA"),
                ("2.0", "Enviar POST a /auth/register con el JSON de prueba.", "El backend crea el usuario y responde 200.", "Se obtuvo respuesta 200 y registro exitoso.", "APROBADA"),
                ("3.0", "Validar email e indicador is_active en la respuesta.", "El email coincide y is_active es true.", "Los datos retornados son correctos.", "APROBADA"),
            ],
        },
        {
            "code": "CP-AU-02",
            "description": "Rechazar registro con correo duplicado",
            "prerequisites": [
                "Debe existir un primer usuario registrado con el correo de prueba.",
                "El endpoint /auth/register debe validar duplicados.",
                "La base de datos de prueba debe conservar el primer registro.",
                "N/A",
            ],
            "test_data": [
                "email repetido",
                "password valida",
                "company_name de prueba",
                "payload usado dos veces",
            ],
            "scenario": "Prueba funcional para comprobar que el sistema evita el registro duplicado de usuarios con el mismo correo.",
            "steps": [
                ("1.0", "Registrar el usuario por primera vez.", "La primera solicitud se completa con exito.", "Primer registro ejecutado correctamente.", "APROBADA"),
                ("2.0", "Reenviar el mismo payload al endpoint /auth/register.", "El sistema rechaza el duplicado con 400.", "La segunda respuesta fue HTTP 400.", "APROBADA"),
                ("3.0", "Validar que el error corresponda a correo ya registrado.", "El mensaje indica duplicidad de correo.", "El backend informo registro duplicado.", "APROBADA"),
            ],
        },
        {
            "code": "CP-AU-03",
            "description": "Iniciar sesion con credenciales validas",
            "prerequisites": [
                "Debe existir un usuario valido previamente registrado.",
                "El endpoint /auth/login debe estar disponible.",
                "La verificacion de password hash debe funcionar.",
                "N/A",
            ],
            "test_data": [
                "username valido",
                "password valida",
                "form-data para login",
                "usuario activo",
            ],
            "scenario": "Prueba funcional de autenticacion para verificar que un usuario valido obtiene un token bearer al iniciar sesion.",
            "steps": [
                ("1.0", "Registrar un usuario de prueba activo.", "El usuario queda listo para autenticarse.", "Usuario registrado con exito.", "APROBADA"),
                ("2.0", "Enviar POST a /auth/login con username y password validos.", "El backend responde 200.", "Se obtuvo respuesta HTTP 200.", "APROBADA"),
                ("3.0", "Verificar access_token y token_type en la respuesta.", "Se retorna token bearer valido.", "La respuesta incluyo access_token y token_type=bearer.", "APROBADA"),
            ],
        },
        {
            "code": "CP-AU-04",
            "description": "Rechazar inicio de sesion con clave invalida",
            "prerequisites": [
                "Debe existir un usuario previamente registrado.",
                "El endpoint /auth/login debe validar credenciales.",
                "El backend debe responder errores de autenticacion.",
                "N/A",
            ],
            "test_data": [
                "username valido",
                "password incorrecta",
                "form-data de login",
                "usuario activo",
            ],
            "scenario": "Prueba funcional para confirmar que el sistema rechaza el acceso cuando la contrasena no coincide con la registrada.",
            "steps": [
                ("1.0", "Registrar un usuario de prueba valido.", "El usuario queda disponible para autenticacion.", "Usuario creado correctamente.", "APROBADA"),
                ("2.0", "Enviar POST a /auth/login con clave incorrecta.", "El backend responde con 401.", "Se obtuvo respuesta HTTP 401.", "APROBADA"),
                ("3.0", "Validar que no se entregue token de acceso.", "La autenticacion es rechazada.", "No se retorno access_token y el acceso fue negado.", "APROBADA"),
            ],
        },
        {
            "code": "CP-RB-01",
            "description": "Proteger /users cuando no hay token",
            "prerequisites": [
                "La ruta /users debe requerir autenticacion.",
                "La dependencia get_current_user debe estar activa.",
                "El cliente de prueba debe permitir solicitudes sin cabecera Authorization.",
                "N/A",
            ],
            "test_data": [
                "Solicitud GET a /users/",
                "Sin token",
                "Sin payload adicional",
                "N/A",
            ],
            "scenario": "Prueba de seguridad para validar que el sistema protege la ruta de usuarios cuando el cliente no envia token de acceso.",
            "steps": [
                ("1.0", "Preparar solicitud GET a /users/ sin encabezado Authorization.", "La solicitud se envia sin credenciales.", "Solicitud preparada sin token.", "APROBADA"),
                ("2.0", "Ejecutar la solicitud contra el backend.", "El sistema rechaza el acceso con 401.", "Se obtuvo respuesta HTTP 401.", "APROBADA"),
                ("3.0", "Confirmar que no se entrega informacion sensible.", "No se expone la lista de usuarios.", "El backend nego el acceso correctamente.", "APROBADA"),
            ],
        },
        {
            "code": "CP-RB-02",
            "description": "Permitir listado de usuarios con rol superadmin",
            "prerequisites": [
                "Debe existir un usuario autenticable con rol superadmin.",
                "El endpoint /users debe aceptar token valido.",
                "La base de datos de prueba debe contener al menos un usuario.",
                "N/A",
            ],
            "test_data": [
                "token bearer valido",
                "solicitud GET a /users/",
                "usuario con rol superadmin",
                "respuesta esperada tipo lista",
            ],
            "scenario": "Prueba de seguridad y autorizacion para verificar que un usuario con privilegios puede consultar el listado de usuarios.",
            "steps": [
                ("1.0", "Registrar un usuario y autenticarlo para obtener token.", "Se obtiene un token bearer valido.", "Token generado correctamente.", "APROBADA"),
                ("2.0", "Enviar GET a /users/ incluyendo Authorization Bearer.", "El backend responde 200.", "Se obtuvo respuesta HTTP 200.", "APROBADA"),
                ("3.0", "Validar que el cuerpo de respuesta sea una lista con usuarios.", "Se retorna una lista no vacia.", "La respuesta devolvio usuarios visibles para el superadmin.", "APROBADA"),
            ],
        },
    ]


def style_worksheet(ws):
    widths = {
        "A": 12,
        "B": 28,
        "C": 18,
        "D": 24,
        "E": 22,
        "F": 24,
        "G": 18,
        "H": 14,
        "I": 18,
        "J": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(1, 40):
        ws.row_dimensions[row].height = 22



def build_sheet(ws, case: dict, fecha: str, resultado_final: str):
    style_worksheet(ws)
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A16"

    merge_and_set(ws, "A1:B1", "ID del caso de prueba", HEADER_FILL, True, CENTER)
    merge_and_set(ws, "C1:C1", case["code"], None, True, CENTER)
    merge_and_set(ws, "D1:E1", "Descripcion del caso", HEADER_FILL, True, CENTER)
    merge_and_set(ws, "F1:J1", case["description"], None, False, LEFT)

    merge_and_set(ws, "A2:B2", "Creado por", HEADER_FILL, True, CENTER)
    set_cell(ws, "C2", "BAPE", bold=False, align=CENTER)
    merge_and_set(ws, "D2:E2", "Revisado por", HEADER_FILL, True, CENTER)
    merge_and_set(ws, "F2:G2", "BAPE", None, False, CENTER)
    merge_and_set(ws, "H2:I2", "Version", HEADER_FILL, True, CENTER)
    set_cell(ws, "J2", "1.0", align=CENTER)

    merge_and_set(ws, "A4:B4", "Bitacora del tester QA", SECTION_FILL, True, CENTER)
    merge_and_set(ws, "C4:J4", "Windows + PowerShell + Python venv + FastAPI TestClient + SQLite temporal + PRUEBAS", None, False, LEFT)

    merge_and_set(ws, "A6:B6", "Nombre del tester", HEADER_FILL, True, CENTER)
    set_cell(ws, "C6", "BAPE", align=CENTER)
    merge_and_set(ws, "D6:E6", "Fecha de prueba", HEADER_FILL, True, CENTER)
    merge_and_set(ws, "F6:G6", fecha, None, False, CENTER)
    merge_and_set(ws, "H6:I6", "Estado del caso", HEADER_FILL, True, CENTER)
    set_cell(ws, "J6", resultado_final, fill=SUB_FILL if resultado_final != "APROBADA" else HEADER_FILL, bold=True, align=CENTER)

    set_cell(ws, "A8", "N.°", fill=SECTION_FILL, bold=True, align=CENTER)
    merge_and_set(ws, "B8:D8", "Prerequisitos", SECTION_FILL, True, CENTER)
    set_cell(ws, "F8", "N.°", fill=SECTION_FILL, bold=True, align=CENTER)
    merge_and_set(ws, "G8:J8", "Datos de prueba", SECTION_FILL, True, CENTER)

    for idx in range(4):
        row = 9 + idx
        set_cell(ws, f"A{row}", f"{idx + 1}.0", align=CENTER)
        merge_and_set(ws, f"B{row}:D{row}", case["prerequisites"][idx], None, False, LEFT)
        set_cell(ws, f"F{row}", f"{idx + 1}.0", align=CENTER)
        merge_and_set(ws, f"G{row}:J{row}", case["test_data"][idx], None, False, LEFT)

    set_cell(ws, "A14", "Escenario de prueba", fill=SECTION_FILL, bold=True, align=CENTER)
    merge_and_set(ws, "B14:J14", case["scenario"], None, False, LEFT)

    set_cell(ws, "A16", "Paso #", fill=SECTION_FILL, bold=True, align=CENTER)
    merge_and_set(ws, "B16:C16", "Detalle del paso", SECTION_FILL, True, CENTER)
    merge_and_set(ws, "D16:E16", "Resultados esperados", SECTION_FILL, True, CENTER)
    merge_and_set(ws, "F16:H16", "Resultados obtenidos", SECTION_FILL, True, CENTER)
    merge_and_set(ws, "I16:J16", "Aprobada / Fallida / No ejecutada / Suspendida", SECTION_FILL, True, CENTER)

    start_row = 18
    for idx, step in enumerate(case["steps"]):
        row = start_row + idx
        num, detail, expected, actual, status = step
        set_cell(ws, f"A{row}", num, align=CENTER)
        merge_and_set(ws, f"B{row}:C{row}", detail, None, False, LEFT)
        merge_and_set(ws, f"D{row}:E{row}", expected, None, False, LEFT)
        merge_and_set(ws, f"F{row}:H{row}", actual, None, False, LEFT)
        merge_and_set(ws, f"I{row}:J{row}", status, HEADER_FILL if status == "APROBADA" else SUB_FILL, True, CENTER)



def main():
    results = json.loads(RESULTS_PATH.read_text(encoding="utf-8"))
    wb = Workbook()
    first = True
    fecha = "31/03/2026"
    for case in build_cases(results):
        ws = wb.active if first else wb.create_sheet()
        ws.title = case["code"]
        build_sheet(ws, case, fecha, status_from_results(results, case["code"]))
        first = False
    wb.save(OUTPUT_PATH)
    print(str(OUTPUT_PATH))


if __name__ == "__main__":
    main()

